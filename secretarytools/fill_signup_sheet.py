#!/usr/bin/env python3
"""
佛學講座簽到單填寫工具

用法:
    python fill_signup_sheet.py <template.xlsx> <output.xlsx> <姓名清單> <講座名稱> <分會名稱>

參數:
    template.xlsx  原始模板檔案
    output.xlsx    輸出檔案路徑
    姓名清單        用逗點隔開，例如: 張三,李四,王五
    講座名稱        帶入 {{Title}}
    分會名稱        帶入 {{Branch}}

範例:
    python fill_signup_sheet.py 佛學講座簽到單.xlsx output.xlsx "張三,李四,王五" "初級禪修課程" "台北分會"
"""

import sys, copy, os, zipfile, shutil, re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

THIN     = Side(border_style="thin")
ALL_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FONT  = Font(name="楷體-簡 標準體", size=16, bold=True)
NUM_FONT     = Font(name="楷體-簡 標準體", size=16, bold=False)
NAME_FONT    = Font(name="楷體-簡 標準體", size=24, bold=False)
SIGN_FONT    = Font(name="楷體-簡 標準體", size=12, bold=False)
CENTER       = Alignment(horizontal="center", vertical="center")
ROW_H_HEADER = 45.0
ROW_H_DATA   = 45.0
COL_WIDTHS   = {"A": 7.5, "B": 17.33, "C": 16.5, "D": 8.33, "E": 19.0, "F": 20.33}


def apply_cell(cell, value, font, alignment=CENTER, border=ALL_THIN):
    cell.value = value
    cell.font = copy.copy(font)
    cell.alignment = copy.copy(alignment)
    cell.border = copy.copy(border)


def set_header_row(ws, row=2):
    for col, h in enumerate(["編號", "姓名", "簽到", "編號", "姓名", "簽到"], start=1):
        apply_cell(ws.cell(row, col), h, HEADER_FONT)
    ws.row_dimensions[row].height = ROW_H_HEADER


def set_data_row(ws, excel_row, left_num, right_num, left_name=None, right_name=None):
    apply_cell(ws.cell(excel_row, 1), left_num,   NUM_FONT)
    apply_cell(ws.cell(excel_row, 2), left_name,  NAME_FONT)
    apply_cell(ws.cell(excel_row, 3), None,        SIGN_FONT)
    apply_cell(ws.cell(excel_row, 4), right_num,  NUM_FONT)
    apply_cell(ws.cell(excel_row, 5), right_name, NAME_FONT)
    apply_cell(ws.cell(excel_row, 6), None,        SIGN_FONT)
    ws.row_dimensions[excel_row].height = ROW_H_DATA


def set_col_widths(ws):
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def fill_and_save(template_path, output_path, names, title, branch):
    n = len(names)

    # ── 計算格數 ──────────────────────────────────
    if n <= 20:
        total = 22
    else:
        candidate = n + 2 if (n + 2) % 2 == 0 else n + 3
        total = candidate
    half = total // 2

    # ── openpyxl: 填入姓名、調整列數、補框線 ────────
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["工作表"]

    # 清除舊資料列
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)

    # 重建合併（openpyxl 需要先 unmerge 再 merge）
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 71.25
    # A1 的值先設成空字串，後面 zip patch 會還原 rich text
    ws["A1"].value = "__TITLE_PLACEHOLDER__"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    set_header_row(ws, row=2)
    set_col_widths(ws)

    for i in range(half):
        left_num   = i + 1
        right_num  = half + i + 1
        left_name  = names[left_num - 1]  if left_num  <= n else None
        right_name = names[right_num - 1] if right_num <= n else None
        actual_right_num = right_num if right_num <= total else None
        set_data_row(ws, i + 3, left_num, actual_right_num, left_name, right_name)

    wb.save(output_path)

    # ── ZIP patch: 還原 A1 rich text，注入 sharedStrings.xml ──
    _patch_title_rich_text(template_path, output_path, title, branch)


def _patch_title_rich_text(template_path, output_path, title, branch):
    """
    openpyxl 在存檔時會把 rich text shared string 轉成 plain inlineStr，
    導致顏色格式丟失。這裡直接操作 ZIP：
    1. 從模板取出 sharedStrings.xml，替換 {{Title}}/{{Branch}}，注入輸出檔
    2. 把 sheet1.xml 裡 A1 的 inlineStr 改回 shared string 參照（s="17" t="s" <v>3</v>）
    3. 更新 [Content_Types].xml 加入 sharedStrings 的 content type（若不存在）
    4. 更新 workbook.xml.rels 加入 sharedStrings 關聯（若不存在）
    """
    # 讀模板的 sharedStrings，替換 placeholder
    with zipfile.ZipFile(template_path, "r") as zt:
        orig_ss   = zt.read("xl/sharedStrings.xml").decode("utf-8")
        orig_rels = zt.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        orig_ct   = zt.read("[Content_Types].xml").decode("utf-8")

    new_ss = orig_ss.replace("{{Title}}", title).replace("{{Branch}}", branch)

    tmp = output_path + ".patching"
    with zipfile.ZipFile(output_path, "r") as zin, \
         zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:

        existing = set(zin.namelist())

        for item in zin.infolist():
            data = zin.read(item.filename)

            if item.filename == "xl/worksheets/sheet1.xml":
                text = data.decode("utf-8")
                # Replace A1 inlineStr back to shared string reference
                # openpyxl uses style index that may differ; keep whatever s= is there
                text = re.sub(
                    r'<c r="A1"([^>]*)>.*?</c>',
                    lambda m: f'<c r="A1"{_keep_style(m.group(1))} t="s"><v>3</v></c>',
                    text, flags=re.DOTALL
                )
                data = text.encode("utf-8")

            elif item.filename == "[Content_Types].xml":
                # Ensure sharedStrings content type is present
                ct = data.decode("utf-8")
                if "sharedStrings" not in ct:
                    ct = ct.replace(
                        "</Types>",
                        '<Override PartName="/xl/sharedStrings.xml" '
                        'ContentType="application/vnd.openxmlformats-officedocument.'
                        'spreadsheetml.sharedStrings+xml"/></Types>'
                    )
                data = ct.encode("utf-8")

            elif item.filename == "xl/_rels/workbook.xml.rels":
                # Ensure sharedStrings relationship is present
                rels = data.decode("utf-8")
                if "sharedStrings" not in rels:
                    rels = rels.replace(
                        "</Relationships>",
                        '<Relationship Id="rIdSS" Type="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships/sharedStrings" '
                        'Target="sharedStrings.xml"/></Relationships>'
                    )
                data = rels.encode("utf-8")

            zout.writestr(item, data)

        # Inject sharedStrings.xml (not written by openpyxl)
        zout.writestr("xl/sharedStrings.xml", new_ss.encode("utf-8"))

    os.replace(tmp, output_path)


def _keep_style(attr_str):
    """從 openpyxl 寫入的屬性字串取出 s= 值，維持樣式索引不變。"""
    m = re.search(r'\bs="(\d+)"', attr_str)
    if m:
        return f' s="{m.group(1)}"'
    return ' s="17"'  # fallback to original template style


def main():
    if len(sys.argv) != 6:
        print(__doc__)
        sys.exit(1)

    template_path = sys.argv[1]
    output_path   = sys.argv[2]
    names_str     = sys.argv[3]
    title         = sys.argv[4]
    branch        = sys.argv[5]

    names = [n.strip() for n in names_str.split(",") if n.strip()]
    if not names:
        print("錯誤：姓名清單不可為空")
        sys.exit(1)

    if "工作表" not in openpyxl.load_workbook(template_path).sheetnames:
        print("錯誤：找不到「工作表」sheet")
        sys.exit(1)

    fill_and_save(template_path, output_path, names, title, branch)
    print(f"✅ 已儲存：{output_path}（{len(names)} 人）")


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        main()
