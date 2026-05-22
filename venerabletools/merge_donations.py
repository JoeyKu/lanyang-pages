#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
捐款名單整合工具
可讀取多個以年份開頭的 Excel 檔案（例如 2024-xxxx.xlsx, 2025-yyyy.xlsx），
將其捐款資料整合，並產生一個包含各年度捐款總額與累計總金額的整合 Excel 檔案。

使用方式：
  1. 自動掃描目前目錄：
     python3 merge_donations.py
  2. 手動指定檔案：
     python3 merge_donations.py 2024-捐款名單.xlsx 2025-捐款名單.xlsx -o 整合結果.xlsx
"""

import os
import sys
import re
import argparse
from typing import Dict, Set, List, Optional, Tuple

# -----------------------------------------------------------------------------
# 自動檢查與安裝依賴套件
# -----------------------------------------------------------------------------
def check_dependencies():
    try:
        import pandas
        import openpyxl
    except ImportError:
        print("💡 偵測到尚未安裝必要套件：pandas 或 openpyxl。")
        print("正在嘗試為您自動安裝...")
        import subprocess
        try:
            subprocess.run([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"], check=True)
            print("✅ 套件安裝成功！繼續執行程式...\n")
        except Exception as e:
            print(f"❌ 自動安裝失敗，錯誤原因: {e}")
            print("\n請在終端機手動執行以下指令安裝依賴套件後，再重新運行此程式：")
            print(f"  {sys.executable} -m pip install pandas openpyxl")
            sys.exit(1)

# 執行套件檢查
check_dependencies()

import pandas as pd

# -----------------------------------------------------------------------------
# 資料結構定義
# -----------------------------------------------------------------------------
class Donor:
    def __init__(self):
        self.names: Set[str] = set()
        self.codes: Set[str] = set()
        self.ids: Set[str] = set()
        self.years: Dict[str, float] = {}  # year -> amount

    def get_display_name(self) -> str:
        """取得最長的名字作為代表名（通常長度較長者資訊較完整）"""
        if not self.names:
            return ""
        return max(self.names, key=len)

    def get_display_code(self) -> str:
        """取得最長的代號作為代表"""
        if not self.codes:
            return ""
        return max(self.codes, key=len)

    def get_display_id(self) -> str:
        """取得身分證號/統編"""
        if not self.ids:
            return ""
        return next(iter(self.ids))

    def update(self, name: str, code: str, id_no: str, year: str, amount: float):
        if name:
            self.names.add(name)
        if code:
            self.codes.add(code)
        if id_no:
            self.ids.add(id_no)
        if year:
            self.years[year] = self.years.get(year, 0.0) + amount

# -----------------------------------------------------------------------------
# 輔助函式
# -----------------------------------------------------------------------------
def get_year_from_filename(filename: str) -> Optional[str]:
    """從檔案名稱中提取 4 位數年份 (如 2025-xxxx.xlsx -> 2025)"""
    basename = os.path.basename(filename)
    # 優先匹配開頭的 4 位數字
    match = re.match(r'^(\d{4})', basename)
    if match:
        return match.group(1)
    # 若開頭沒有，則在檔名中任意搜尋 4 位數字
    match_any = re.search(r'(\d{4})', basename)
    if match_any:
        return match_any.group(1)
    return None

def clean_to_string(val) -> str:
    """清理 Excel 讀入的資料為乾淨的字串，並處理浮點數 (如 123.0 -> "123")"""
    if pd.isna(val):
        return ""
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val).strip()
    return str(val).strip()

def read_excel_with_header_detection(filepath: str) -> Tuple[pd.DataFrame, int]:
    """自動偵測標題列並讀取 Excel 檔案"""
    # 先以無標題方式讀取前幾列，以便尋找標題關鍵字所在行數
    df_preview = pd.read_excel(filepath, header=None, nrows=15)
    
    header_row_idx = 0
    found = False
    
    for idx, row in df_preview.iterrows():
        # 將該列的所有非空值轉成字串
        row_str = [str(x).strip() for x in row.values if pd.notna(x)]
        # 檢查該列是否同時包含「姓名/捐款人」以及「金額/總計」相關關鍵字
        has_name = any(any(k in s for k in ['捐款人', '姓名']) for s in row_str)
        has_amount = any(any(k in s for k in ['金額', '總金額', '總計', '累計', '金額小計']) for s in row_str)
        
        if has_name and has_amount:
            header_row_idx = idx
            found = True
            break
            
    if found:
        print(f"   🔍 偵測到標題列位於第 {header_row_idx + 1} 列，以此行開始讀取。")
        return pd.read_excel(filepath, skiprows=header_row_idx), header_row_idx
    else:
        print("   ⚠️ 未偵測到明確的欄位標題特徵，採用預設首列作為欄位名稱。")
        return pd.read_excel(filepath), 0

def find_matched_columns(df: pd.DataFrame) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
    """根據列名比對並對應出所需要的四個欄位"""
    # 清理所有欄位名稱便於匹配
    columns = [str(col).strip() for col in df.columns]
    original_cols = list(df.columns)
    
    # --- 1. 總金額欄位定位 ---
    # 嚴格精準匹配「總金額」，若無此欄位則不讀取
    amount_col = None
    for cleaned_name, orig_name in zip(columns, original_cols):
        if cleaned_name == '總金額':
            amount_col = orig_name
            break

    # --- 2. 捐款人姓名定位 ---
    # 第一優先：精準匹配「捐款人」
    name_col = None
    for cleaned_name, orig_name in zip(columns, original_cols):
        if cleaned_name == '捐款人':
            name_col = orig_name
            break
    # 第二優先：精準匹配「姓名」
    if not name_col:
        for cleaned_name, orig_name in zip(columns, original_cols):
            if cleaned_name == '姓名':
                name_col = orig_name
                break
    # 第三優先：包含「捐款人」或「姓名」
    if not name_col:
        for cleaned_name, orig_name in zip(columns, original_cols):
            if any(k in cleaned_name for k in ['捐款人', '姓名']):
                if not any(ex in cleaned_name for ex in ['代號', '編號', '代碼', '統編', '身分證']):
                    name_col = orig_name
                    break

    # --- 3. 捐款人代號定位 ---
    # 第一優先：精準匹配「捐款人代號」或「捐款人編號」
    code_col = None
    for cleaned_name, orig_name in zip(columns, original_cols):
        if cleaned_name in ['捐款人代號', '捐款人編號']:
            code_col = orig_name
            break
    # 第二優先：包含「代號」、「編號」且排除身分證/金額相關
    if not code_col:
        for cleaned_name, orig_name in zip(columns, original_cols):
            if any(k in cleaned_name for k in ['代號', '編號', '代碼', '會員編號']):
                if not any(ex in cleaned_name for ex in ['身分證', '統編', '金額', '證號']):
                    code_col = orig_name
                    break

    # --- 4. 身分證/統編定位 ---
    # 第一優先：精準匹配「身分證/統編」或「身分證字號」或「統一編號」
    id_col = None
    for cleaned_name, orig_name in zip(columns, original_cols):
        if cleaned_name in ['身分證/統編', '身分證字號', '統一編號']:
            id_col = orig_name
            break
    # 第二優先：包含「身分證」、「統編」、「證號」
    if not id_col:
        for cleaned_name, orig_name in zip(columns, original_cols):
            if any(k in cleaned_name for k in ['身分證', '統編', '證號']):
                id_col = orig_name
                break
                
    return name_col, code_col, id_col, amount_col

def find_matching_donor(donors: List[Donor], name: str, code: str, id_no: str) -> Optional[Donor]:
    """比對現有的捐款人資料以進行整合，防止不同人因重名或空值被錯誤合併"""
    # 1. 優先使用「捐款人代號」匹配
    if code:
        for d in donors:
            if code in d.codes:
                return d
                
    # 2. 其次使用「身分證/統編」匹配 (確保無代號衝突)
    if id_no:
        for d in donors:
            if id_no in d.ids:
                # 若此筆資料有代號，且現有捐款人也有不同代號，則屬於衝突，不合併
                if code and d.codes and code not in d.codes:
                    continue
                return d
                
    # 3. 最後使用「姓名」匹配 (僅在兩者均無 Code 與 ID 衝突時合併，避免同名同姓誤判)
    if name:
        for d in donors:
            if name in d.names:
                # 排除任何 Code 或 ID 的衝突
                if code and d.codes and code not in d.codes:
                    continue
                if id_no and d.ids and id_no not in d.ids:
                    continue
                return d
                
    return None

# -----------------------------------------------------------------------------
# 主程式
# -----------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="將多個年度的捐款 Excel 檔案整合為一筆並產出新檔。")
    parser.add_argument("files", nargs="*", help="要整合的 Excel 檔案路徑（如未指定，將搜尋當前目錄）")
    parser.add_argument("-o", "--output", default="整合捐款名單.xlsx", help="輸出之整合 Excel 檔案名稱 (預設: 整合捐款名單.xlsx)")
    parser.add_argument("-d", "--dir", default=".", help="未指定檔案時，要掃描的目錄 (預設: 當前目錄)")
    args = parser.parse_args()

    files_to_process = args.files
    
    # 若無指定檔案，自動搜尋指定目錄下符合 \d{4}- 格式的 Excel
    if not files_to_process:
        print(f"📂 正在掃描目錄 '{args.dir}' 中的 Excel 檔案...")
        all_files = os.listdir(args.dir)
        for f in all_files:
            if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$'):
                year = get_year_from_filename(f)
                if year:
                    files_to_process.append(os.path.join(args.dir, f))
                    
        if not files_to_process:
            print("❌ 找不到以年份開頭的 Excel 檔案（例如 2025-xxxx.xlsx）。")
            print("請手動將檔案拖入此目錄，或使用引數指定檔案路徑。")
            return
            
        print(f"✅ 找到 {len(files_to_process)} 個待處理的 Excel 檔案：")
        for idx, filepath in enumerate(files_to_process, 1):
            year = get_year_from_filename(filepath)
            print(f"  {idx}. {os.path.basename(filepath)} (解析年份: {year})")
            
        # 互動確認
        try:
            confirm = input("\n是否開始進行整合？ [Y/n]: ").strip().lower()
            if confirm not in ['', 'y', 'yes']:
                print("已取消操作。")
                return
        except KeyboardInterrupt:
            print("\n已取消操作。")
            return

    # 開始解析與整合
    donors: List[Donor] = []
    all_years: Set[str] = set()
    
    print("\n🚀 開始處理檔案...")
    for filepath in files_to_process:
        if not os.path.exists(filepath):
            print(f"❌ 檔案不存在: {filepath}，跳過此檔案。")
            continue
            
        filename = os.path.basename(filepath)
        year = get_year_from_filename(filename)
        
        if not year:
            print(f"⚠️ 無法從檔名 '{filename}' 解析出年份，請輸入該檔案的年份（直接按 Enter 鍵跳過此檔）:")
            user_year = input("年份: ").strip()
            if not user_year:
                print(f"⏭️ 已跳過檔案: {filename}")
                continue
            year = user_year
            
        all_years.add(year)
        print(f"\n📖 讀取檔案: {filename} (歸類為 {year} 年度)")
        
        try:
            df, header_row_idx = read_excel_with_header_detection(filepath)
        except Exception as e:
            print(f"❌ 讀取 Excel 失敗: {e}")
            continue
            
        name_col, code_col, id_col, amount_col = find_matched_columns(df)
        
        # 列出比對到的欄位
        print(f"   對應欄位結果:")
        print(f"     - 捐款人姓名: {f'[{name_col}]' if name_col else '❌ 未找到'}")
        print(f"     - 捐款人代號: {f'[{code_col}]' if code_col else '❌ 未找到'}")
        print(f"     - 身分證/統編: {f'[{id_col}]' if id_col else '⚠️ 未找到 (選填)'}")
        print(f"     - 總金額欄位: {f'[{amount_col}]' if amount_col else '❌ 未找到'}")
        
        if not name_col or not amount_col or not code_col:
            print(f"\n❌ 關鍵錯誤：檔案 '{filename}' 缺少必要欄位！")
            if not name_col:
                print("   - 缺少 [捐款人] 或 [姓名] 欄位")
            if not code_col:
                print("   - 缺少 [捐款人代號] 欄位")
            if not amount_col:
                print("   - 缺少 [總金額] 欄位")
            print("🛑 整合程序已終止，未產生任何整合檔案。")
            sys.exit(1)
            
        # 逐筆載入資料
        success_count = 0
        empty_row_count = 0
        abnormal_row_count = 0
        
        for idx, row in df.iterrows():
            # 計算 Excel 的真實列號 (0-indexed index + header offset + 2)
            excel_row = idx + header_row_idx + 2
            
            # 讀取欄位資料並清理
            name = clean_to_string(row[name_col])
            code = clean_to_string(row[code_col]) if code_col else ""
            id_no = clean_to_string(row[id_col]) if id_col else ""
            
            # 若為身分證，轉為大寫
            if id_no:
                id_no = id_no.upper()
                
            # 金額解析
            raw_amount = row[amount_col]
            amount = 0.0
            amount_parsed = False
            
            if pd.notna(raw_amount) and str(raw_amount).strip() != "":
                try:
                    # 移除非數字字元 (例如千分位逗號, 錢字符號)
                    amount_str = re.sub(r'[^\d\.\-]', '', str(raw_amount))
                    if amount_str:
                        amount = float(amount_str)
                        amount_parsed = True
                    else:
                        amount = 0.0
                        amount_parsed = True
                except ValueError:
                    pass
            else:
                amount = 0.0
                amount_parsed = True
                
            # 判斷是否為完全空白列 (姓名、代號、身分證、原始金額皆無內容)
            is_empty_row = (not name) and (not code) and (not id_no) and (pd.isna(raw_amount) or str(raw_amount).strip() == "")
            
            if is_empty_row:
                empty_row_count += 1
                continue
                
            # 關鍵檢查：任何有效資料列，若缺少姓名或代號，直接報錯並終止執行
            if not name or not code:
                print(f"\n❌ 關鍵錯誤：檔案 '{filename}' 第 {excel_row} 列缺少必要內容！")
                if not name:
                    print("   - 缺少「捐款人」姓名")
                if not code:
                    print("   - 缺少「捐款人代號」")
                print("🛑 整合程序已終止，未產生任何整合檔案。")
                sys.exit(1)
                
            # 若金額無法解析
            if not amount_parsed:
                print(f"     ⚠️ 第 {excel_row} 列金額格式異常：原始內容為「{raw_amount}」，無法解析為數值，已自動將金額設為 0 元載入。")
                
            # 比對現有捐款人
            donor = find_matching_donor(donors, name, code, id_no)
            if not donor:
                donor = Donor()
                donors.append(donor)
                
            donor.update(name=name, code=code, id_no=id_no, year=year, amount=amount)
            success_count += 1
            
        print(f"   📊 檔案處理完成：成功讀取 {success_count} 筆資料 (過濾空白列 {empty_row_count} 筆, 異常跳過 {abnormal_row_count} 筆)")

    if not donors:
        print("\n❌ 整合完成，但沒有讀取到任何有效資料。")
        return

    # 排序年份，使欄位輸出時依年份先後排序
    sorted_years = sorted(list(all_years))
    
    # 準備整理為 DataFrame 輸出
    output_rows = []
    for d in donors:
        row_dict = {
            '捐款人': d.get_display_name(),
            '捐款人代號': d.get_display_code(),
            '身分證/統編': d.get_display_id(),
        }
        # 各年份金額，若無則為 0
        for y in sorted_years:
            row_dict[f'總金額-{y}'] = d.years.get(y, 0.0)
            
        # 計算加總總金額
        row_dict['總金額'] = sum(d.years.values())
        output_rows.append(row_dict)
        
    df_output = pd.DataFrame(output_rows)
    
    # 格式化輸出 Excel (使欄位排序漂亮)
    cols_order = ['捐款人', '捐款人代號', '身分證/統編'] + [f'總金額-{y}' for y in sorted_years] + ['總金額']
    # 過濾只存在於 dataframe 中的欄位，防止報錯
    cols_order = [c for c in cols_order if c in df_output.columns]
    df_output = df_output[cols_order]
    
    # 根據總金額降序排序
    if '總金額' in df_output.columns:
        df_output = df_output.sort_values(by='總金額', ascending=False)
    
    # 寫入 Excel
    output_filename = args.output
    try:
        # 使用 openpyxl 引擎寫入，以利格式最佳化
        writer = pd.ExcelWriter(output_filename, engine='openpyxl')
        df_output.to_excel(writer, index=False, sheet_name='捐款名單整合')
        
        workbook = writer.book
        worksheet = writer.sheets['捐款名單整合']
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # 配色定義 (琥珀金棕/暖心主題)
        primary_color = "A87A4F"    # 標題列背景 (深暖棕)
        zebra_color = "FDFAF3"      # 偶數列背景 (極淺暖金黃)
        total_col_color = "F5EBE0"  # 總金額整欄背景 (溫潤米色)
        border_color = "E6DFD5"     # 網格線顏色 (淺灰褐)
        
        # 字型設定 (預設微軟正黑體)
        font_name = "Microsoft JhengHei"
        header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_name, size=11, bold=False, color="3D2C1B")
        total_font = Font(name=font_name, size=11, bold=True, color="3D2C1B")
        
        # 填充樣式
        header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
        zebra_fill = PatternFill(start_color=zebra_color, end_color=zebra_color, fill_type="solid")
        total_col_fill = PatternFill(start_color=total_col_color, end_color=total_col_color, fill_type="solid")
        
        # 邊框樣式
        thin_side = Side(style='thin', color=border_color)
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # 對齊方式
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # 標題列美化
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        worksheet.row_dimensions[1].height = 28  # 標題列高度加寬，較為大氣
        
        max_col = worksheet.max_column
        
        # 資料列美化與格式設定
        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = 22  # 資料列高度加寬，便於閱讀
            is_even = (row_idx % 2 == 0)
            
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                # 1. 設置背景填充與字型 (總金額欄特別凸顯，其餘偶數列採斑馬紋)
                if col_idx == max_col:
                    cell.fill = total_col_fill
                    cell.font = total_font
                elif is_even:
                    cell.fill = zebra_fill
                
                # 2. 設置對齊方式
                if col_idx in [1, 3]:  # 捐款人、身分證/統編
                    cell.alignment = left_align
                elif col_idx == 2:  # 捐款人代號
                    cell.alignment = center_align
                else:  # 年份金額與總金額
                    cell.alignment = right_align
                    
                # 3. 設置千分位數值格式 (若為金額欄位)
                if col_idx >= 4:
                    cell.number_format = '#,##0'
                    
        # 表格欄寬自適應調整 (考慮中文字寬度)
        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or '')
                # 計算中文字元與英文字元的實際字寬
                cell_len = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in val_str)
                if cell_len > max_len:
                    max_len = cell_len
            # 設定合適的最小與最大寬度
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        writer.close()
        print(f"\n🎉 整合成功！")
        print(f"📁 輸出檔案：{os.path.abspath(output_filename)}")
        print(f"👥 整合後不重複捐款人總數：{len(df_output)} 筆")
    except Exception as e:
        print(f"\n❌ 儲存整合 Excel 失敗: {e}")

if __name__ == "__main__":
    main()
